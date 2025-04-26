import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from openpyxl import load_workbook, Workbook
from PIL import Image, ImageTk
import random
import os
import threading
import time
import sys
from datetime import datetime

class LotteryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("سامانه قرعه‌کشی هوشمند")
        self.entries = []
        self.bg_path = None
        self.default_bg = self.resource_path("default_bg.jpg")
        self.countdown_seconds = 5
        self.winner_count = 1
        self.is_spinning = False
        self.previous_winners = []
        self.current_theme = "dark"
        
        # Load icon
        try:
            icon_path = self.resource_path("lottery_icon.ico")
            if os.path.exists(icon_path):
                root.iconbitmap(icon_path)
        except Exception as e:
            print(f"Could not load icon: {e}")
        
        # Set default theme
        self.set_theme(self.current_theme)
        
        self.setup_ui()
        self.create_menu()
        
        # Center the window
        self.center_window()

    def resource_path(self, relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)

    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def set_theme(self, theme):
        """Set application theme (dark or light)"""
        self.current_theme = theme
        
        if theme == "dark":
            self.bg_color = '#2d2d2d'
            self.fg_color = 'white'
            self.accent_color = '#28a745'
            self.secondary_color = '#6c757d'
            self.text_bg = '#343a40'
            self.highlight_color = '#ffc107'
            self.button_bg = '#495057'
            self.winner_bg = '#495057'
            self.winner_fg = 'white'
        else:
            self.bg_color = '#f8f9fa'
            self.fg_color = '#212529'
            self.accent_color = '#007bff'
            self.secondary_color = '#6c757d'
            self.text_bg = 'white'
            self.highlight_color = '#fd7e14'
            self.button_bg = '#e9ecef'
            self.winner_bg = '#e9ecef'
            self.winner_fg = '#212529'
        
        # Apply colors to widgets
        self.root.config(bg=self.bg_color)
        if hasattr(self, 'count_label'):
            self.count_label.config(background=self.bg_color, foreground=self.fg_color)
        if hasattr(self, 'status_bar'):
            self.status_bar.config(background=self.secondary_color, foreground='white')

    def create_menu(self):
        menubar = tk.Menu(self.root)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0, bg=self.bg_color, fg=self.fg_color)
        file_menu.add_command(label="باز کردن فایل Excel", command=self.load_excel)
        file_menu.add_command(label="انتخاب تصویر پس‌زمینه", command=self.select_background)
        file_menu.add_separator()
        file_menu.add_command(label="تغییر تم", command=self.toggle_theme)
        file_menu.add_separator()
        file_menu.add_command(label="خروج", command=self.root.quit)
        menubar.add_cascade(label="فایل", menu=file_menu)
        
        # Winners menu
        winner_menu = tk.Menu(menubar, tearoff=0, bg=self.bg_color, fg=self.fg_color)
        winner_menu.add_command(label="مشاهده برندگان قبلی", command=self.show_previous_winners)
        winner_menu.add_command(label="پاک کردن لیست برندگان", command=self.clear_winners)
        winner_menu.add_command(label="ذخیره نتایج به صورت Excel", command=self.save_winners_explicit)
        menubar.add_cascade(label="برندگان", menu=winner_menu)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0, bg=self.bg_color, fg=self.fg_color)
        help_menu.add_command(label="راهنمای استفاده", command=self.show_help)
        help_menu.add_command(label="درباره برنامه", command=self.show_about)
        menubar.add_cascade(label="راهنما", menu=help_menu)
        
        self.root.config(menu=menubar)

    def toggle_theme(self):
        """Toggle between dark and light theme"""
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.set_theme(self.current_theme)
        messagebox.showinfo("تم تغییر کرد", f"تم برنامه به {self.current_theme} تغییر یافت.")

    def setup_ui(self):
        self.root.geometry("900x650")
        self.root.minsize(800, 550)
        
        # General styles
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure styles
        style.configure('.', background=self.bg_color, foreground=self.fg_color)
        style.configure('TFrame', background=self.bg_color)
        style.configure('TLabel', background=self.bg_color, foreground=self.fg_color, font=('B Titr', 12))
        style.configure('TButton', font=('B Titr', 12), padding=8)
        style.configure('Accent.TButton', font=('B Titr', 14), background=self.accent_color, foreground='white')
        style.configure('TEntry', font=('B Titr', 12), padding=5, fieldbackground=self.text_bg)
        
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Application title
        title_label = ttk.Label(
            main_frame, 
            text="سامانه قرعه‌کشی هوشمند", 
            font=("B Titr", 24, "bold"),
            foreground=self.highlight_color
        )
        title_label.pack(pady=(0, 20))
        
        # Frame for buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        # Buttons
        self.excel_btn = ttk.Button(
            button_frame, 
            text="📥 انتخاب فایل Excel", 
            command=self.load_excel,
            style='TButton'
        )
        self.excel_btn.pack(side=tk.RIGHT, padx=5, expand=True)
        
        self.bg_btn = ttk.Button(
            button_frame, 
            text="🖼 انتخاب تصویر پس‌زمینه", 
            command=self.select_background,
            style='TButton'
        )
        self.bg_btn.pack(side=tk.LEFT, padx=5, expand=True)
        
        # Settings frame
        settings_frame = ttk.Frame(main_frame)
        settings_frame.pack(fill=tk.X, pady=10)
        
        # Number of winners
        winner_frame = ttk.Frame(settings_frame)
        winner_frame.pack(side=tk.RIGHT, padx=10, expand=True)
        
        ttk.Label(winner_frame, text="🎯 تعداد برنده‌ها:").pack(anchor=tk.E)
        self.winner_entry = ttk.Entry(winner_frame, justify='center', width=10)
        self.winner_entry.insert(0, "1")
        self.winner_entry.pack(fill=tk.X)
        
        # Countdown duration
        timer_frame = ttk.Frame(settings_frame)
        timer_frame.pack(side=tk.LEFT, padx=10, expand=True)
        
        ttk.Label(timer_frame, text="🕒 مدت شمارش معکوس (ثانیه):").pack(anchor=tk.E)
        self.timer_entry = ttk.Entry(timer_frame, justify='center', width=10)
        self.timer_entry.insert(0, "5")
        self.timer_entry.pack(fill=tk.X)
        
        # Display participant count
        self.count_label = ttk.Label(
            main_frame, 
            text="👥 تعداد افراد: 0", 
            font=("B Titr", 14),
            foreground=self.highlight_color
        )
        self.count_label.pack(pady=20)
        
        # Start button
        self.start_btn = ttk.Button(
            main_frame, 
            text="🚀 شروع قرعه‌کشی", 
            command=self.start_lottery, 
            style='Accent.TButton'
        )
        self.start_btn.pack(pady=20, fill=tk.X)
        
        # Status bar
        self.status_bar = ttk.Label(
            main_frame, 
            text="آماده به کار", 
            relief=tk.SUNKEN, 
            anchor=tk.W,
            background=self.secondary_color,
            foreground='white',
            font=('B Titr', 10)
        )
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def show_help(self):
        help_text = """
        راهنمای استفاده از سامانه قرعه‌کشی:
        
        1. ابتدا فایل Excel حاوی اطلاعات شرکت‌کنندگان را انتخاب کنید.
        2. در صورت تمایل می‌توانید یک تصویر برای پس‌زمینه انتخاب کنید.
        3. تعداد برنده‌ها و مدت زمان شمارش معکوس را مشخص کنید.
        4. دکمه "شروع قرعه‌کشی" را کلیک کنید.
        
        ساختار فایل Excel باید به صورت زیر باشد:
        - ستون اول: نام شرکت‌کننده
        - ستون دوم: کد ملی
        - ستون سوم: شماره موبایل
        
        ویژگی‌های جدید:
        - امکان تغییر تم برنامه
        - نمایش بهتر برنده‌های زیاد
        - ذخیره‌سازی خودکار نتایج
        - جلوگیری از انتخاب تکراری برنده‌ها
        - امکان ذخیره دستی نتایج
        """
        messagebox.showinfo("راهنمای برنامه", help_text)

    def show_about(self):
        about_text = """
        سامانه قرعه‌کشی هوشمند
        
        نسخه 2.1
        
        ویژگی‌ها:
        - رابط کاربری بهبود یافته
        - پشتیبانی از تم‌های مختلف
        - نمایش داینامیک نتایج
        - ذخیره‌سازی خودکار برندگان
        - سازگاری با سیستم‌های مختلف
        
        توسعه داده شده با Python و Tkinter
        © 2023
        """
        messagebox.showinfo("درباره برنامه", about_text)

    def show_previous_winners(self):
        if not self.previous_winners:
            messagebox.showinfo("برندگان", "هنوز برنده‌ای انتخاب نشده است.")
            return
            
        win_text = "\n\n".join([
            f"{i+1}- {name}\nکد ملی: {nid} / تلفن: {self.mask_phone(phone)}"
            for i, (name, nid, phone) in enumerate(self.previous_winners)
        ])
        
        # Create new window for winners
        win_window = tk.Toplevel(self.root)
        win_window.title("لیست برندگان قبلی")
        win_window.geometry("600x400")
        
        # Center the window
        win_window.update_idletasks()
        width = win_window.winfo_width()
        height = win_window.winfo_height()
        x = (win_window.winfo_screenwidth() // 2) - (width // 2)
        y = (win_window.winfo_screenheight() // 2) - (height // 2)
        win_window.geometry(f'+{x}+{y}')
        
        # Use ScrolledText for long lists
        text_area = scrolledtext.ScrolledText(
            win_window,
            wrap=tk.WORD,
            font=('B Titr', 12),
            bg=self.text_bg,
            fg=self.fg_color,
            padx=10,
            pady=10
        )
        text_area.pack(fill=tk.BOTH, expand=True)
        text_area.insert(tk.INSERT, win_text)
        text_area.config(state=tk.DISABLED)
        
        close_btn = ttk.Button(
            win_window,
            text="بستن",
            command=win_window.destroy
        )
        close_btn.pack(pady=10)

    def clear_winners(self):
        if messagebox.askyesno("تأیید", "آیا مطمئن هستید که می‌خواهید لیست برندگان قبلی پاک شود؟"):
            self.previous_winners = []
            messagebox.showinfo("موفق", "لیست برندگان قبلی پاک شد.")
            self.status_bar.config(text="لیست برندگان قبلی پاک شد")

    def save_winners_explicit(self):
        if not self.previous_winners:
            messagebox.showwarning("هشدار", "لیست برندگان خالی است.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="ذخیره نتایج به عنوان"
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Winners"
                ws.append(["ردیف", "نام", "کد ملی", "شماره موبایل (مخفی)", "تاریخ قرعه‌کشی"])
                
                # Set column widths
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                
                for i, (name, national_id, phone) in enumerate(self.previous_winners, start=1):
                    ws.append([
                        i,
                        name,
                        national_id,
                        self.mask_phone(phone),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ])
                
                wb.save(file_path)
                messagebox.showinfo("موفق", f"نتایج با موفقیت در {file_path} ذخیره شد.")
                self.status_bar.config(text=f"نتایج در {os.path.basename(file_path)} ذخیره شد")
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در ذخیره فایل:\n{e}")
                self.status_bar.config(text="خطا در ذخیره نتایج")

    def load_excel(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            title="لطفاً فایل Excel را انتخاب کنید"
        )
        if not file_path:
            return
        
        self.status_bar.config(text="در حال بارگذاری فایل...")
        self.root.update()
        
        try:
            wb = load_workbook(file_path)
            sheet = wb.active
            self.entries = []

            for row in sheet.iter_rows(min_row=1, values_only=True):
                if len(row) >= 3:  # Ensure we have at least 3 columns
                    name, national_id, phone = row[:3]
                    if name and national_id and phone:
                        self.entries.append((str(name).strip(), str(national_id).strip(), str(phone).strip()))

            if not self.entries:
                messagebox.showwarning("هشدار", "فایل انتخاب شده حاوی اطلاعات معتبر نیست.")
                self.status_bar.config(text="فایل حاوی اطلاعات معتبر نیست")
                return

            self.count_label.config(text=f"👥 تعداد افراد: {len(self.entries)}")
            self.status_bar.config(text=f"فایل با موفقیت بارگذاری شد. تعداد شرکت‌کنندگان: {len(self.entries)}")
            messagebox.showinfo("موفق", "اطلاعات با موفقیت بارگذاری شد.")
        except Exception as e:
            messagebox.showerror("خطا", f"در خواندن فایل مشکلی پیش آمد:\n{str(e)}")
            self.status_bar.config(text="خطا در بارگذاری فایل")

    def select_background(self):
        path = filedialog.askopenfilename(
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp"), ("All files", "*.*")],
            title="لطفاً تصویر پس‌زمینه را انتخاب کنید"
        )
        if path:
            try:
                # Verify it's a valid image file
                with Image.open(path) as img:
                    self.bg_path = path
                    messagebox.showinfo("ثبت شد", "تصویر پس‌زمینه تنظیم شد.")
                    self.status_bar.config(text="تصویر پس‌زمینه انتخاب شد")
            except Exception as e:
                messagebox.showerror("خطا", f"فایل انتخاب شده یک تصویر معتبر نیست:\n{e}")
                self.status_bar.config(text="خطا در انتخاب تصویر")

    def start_lottery(self):
        if not self.entries:
            messagebox.showwarning("هشدار", "لطفاً ابتدا فایل Excel را بارگذاری کنید.")
            self.status_bar.config(text="خطا: فایل Excel بارگذاری نشده است")
            return

        try:
            self.winner_count = int(self.winner_entry.get())
            self.countdown_seconds = int(self.timer_entry.get())
            
            if self.winner_count <= 0 or self.countdown_seconds <= 0:
                raise ValueError
                
        except ValueError:
            messagebox.showerror("خطا", "مقادیر وارد شده باید اعداد صحیح مثبت باشند.")
            self.status_bar.config(text="خطا: مقادیر ورودی نامعتبر")
            return

        available_entries = [e for e in self.entries if e not in self.previous_winners]

        if self.winner_count > len(available_entries):
            messagebox.showerror("خطا", f"تعداد برنده‌ها بیشتر از افراد باقیمانده است. فقط {len(available_entries)} شرکت‌کننده باقی مانده.")
            self.status_bar.config(text=f"خطا: فقط {len(available_entries)} شرکت‌کننده باقی مانده")
            return

        self.start_btn.config(state=tk.DISABLED)
        self.status_bar.config(text="در حال آماده‌سازی قرعه‌کشی...")
        threading.Thread(target=self.run_lottery, daemon=True).start()

    def run_lottery(self):
        popup = tk.Toplevel(self.root)
        popup.attributes('-fullscreen', True)
        popup.configure(bg='black')
        popup.title("قرعه‌کشی در حال انجام...")

        # Display background image
        bg_img = None
        try:
            if self.bg_path and os.path.exists(self.bg_path):
                bg_img = Image.open(self.bg_path)
            elif os.path.exists(self.default_bg):
                bg_img = Image.open(self.default_bg)
                
            if bg_img:
                screen_width = popup.winfo_screenwidth()
                screen_height = popup.winfo_screenheight()
                
                # Maintain aspect ratio while resizing
                img_ratio = bg_img.width / bg_img.height
                screen_ratio = screen_width / screen_height
                
                if img_ratio > screen_ratio:
                    # Image is wider than screen
                    new_width = screen_width
                    new_height = int(screen_width / img_ratio)
                else:
                    # Image is taller than screen
                    new_height = screen_height
                    new_width = int(screen_height * img_ratio)
                
                bg_img = bg_img.resize((new_width, new_height), Image.LANCZOS)
                bg_photo = ImageTk.PhotoImage(bg_img)
                bg_label = tk.Label(popup, image=bg_photo)
                bg_label.image = bg_photo
                bg_label.place(x=(screen_width - new_width) // 2, y=(screen_height - new_height) // 2)
        except Exception as e:
            print(f"Error loading background: {e}")

        # Countdown
        countdown_label = tk.Label(
            popup, 
            font=("B Titr", 100), 
            fg="white", 
            bg="black",
            relief=tk.RAISED,
            bd=5
        )
        countdown_label.place(relx=0.5, rely=0.5, anchor="center")

        for i in range(self.countdown_seconds, 0, -1):
            countdown_label.config(text=str(i))
            popup.update()
            time.sleep(1)

        countdown_label.destroy()

        # Spinner display
        spin_label = tk.Label(
            popup, 
            font=("B Titr", 40),
            fg="yellow", 
            bg="black", 
            justify="center",
            relief=tk.RAISED,
            bd=5,
            anchor='center'
        )
        spin_label.place(relx=0.5, rely=0.5, anchor="center")

        self.is_spinning = True
        spinner_thread = threading.Thread(target=self.spin_names, args=(popup, spin_label))
        spinner_thread.start()

        popup.after(self.countdown_seconds * 1000, lambda: self.stop_spinning(popup, spin_label))

        # Control buttons
        control_frame = tk.Frame(popup, bg='black')
        control_frame.place(relx=0.5, rely=0.9, anchor="center")
        
        retry_btn = tk.Button(
            control_frame, 
            text="🔁 قرعه‌کشی مجدد", 
            font=("B Titr", 20), 
            command=lambda: self.retry_lottery(popup, spin_label), 
            bg="#007bff", 
            fg="white",
            padx=20,
            pady=10,
            bd=0,
            relief=tk.RAISED
        )
        retry_btn.pack(side=tk.LEFT, padx=10)
        
        close_btn = tk.Button(
            control_frame, 
            text="✕ بستن", 
            font=("B Titr", 20), 
            command=popup.destroy, 
            bg="#dc3545", 
            fg="white",
            padx=20,
            pady=10,
            bd=0,
            relief=tk.RAISED
        )
        close_btn.pack(side=tk.RIGHT, padx=10)

        popup.bind("<Escape>", lambda e: popup.destroy())

    def spin_names(self, popup, label):
        while self.is_spinning:
            entry = random.choice(self.entries)
            print(self.mask_phone(entry[2]))
            name_text = f"{entry[0]}\n{entry[1]} / {self.mask_phone(entry[2])}"
            label.config(text=name_text)
            popup.update()
            time.sleep(0.1)

    def stop_spinning(self, popup, label):
        self.is_spinning = False

        available_entries = [e for e in self.entries if e not in self.previous_winners]
        winners = random.sample(available_entries, self.winner_count)

        # Create frame for winners with scrollbar
        result_frame = tk.Frame(popup, bg='black')
        result_frame.place(relx=0.5, rely=0.5, anchor="center", relwidth=0.9, relheight=0.7)
        
        canvas = tk.Canvas(result_frame, bg='black', highlightthickness=0)
        scrollbar = tk.Scrollbar(result_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='black')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="center")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Set font size based on number of winners
        if self.winner_count <= 3:
            font_size = 40
        elif self.winner_count <= 6:
            font_size = 30
        elif self.winner_count <= 10:
            font_size = 24
        elif self.winner_count <= 20:
            font_size = 18
        else:
            font_size = 14

        # Create a centered container for winners
        center_frame = tk.Frame(scrollable_frame, bg='black')
        center_frame.pack()
        
        # Display winners in a centered column
        for i, (name, national_id, phone) in enumerate(winners, start=1):
            winner_frame = tk.Frame(
                center_frame, 
                bg=self.winner_bg, 
                bd=2, 
                relief=tk.RAISED,
                padx=20,
                pady=10
            )
            winner_frame.pack(pady=10, anchor='center', expand=True)
            # Winner number
            # tk.Label(
            #     winner_frame,
            #     text=f"برنده شماره {i}",
            #     font=("B Titr", font_size-4),
            #     fg=self.highlight_color,
            #     bg=self.winner_bg,
            #     anchor='center'
            # ).pack(fill=tk.X, expand=True)
            
            # Winner name
            tk.Label(
                winner_frame,
                text=f"🏆 {name}",
                font=("B Titr", font_size),
                fg=self.highlight_color,
                bg=self.winner_bg,
                anchor='center'
            ).pack(fill=tk.X, expand=True)
            
            # National ID
            tk.Label(
                winner_frame,
                text=f"کد ملی: {national_id}",
                font=("B Titr", font_size-4),
                fg=self.winner_fg,
                bg=self.winner_bg,
                anchor='center'
            ).pack(fill=tk.X, expand=True)
            # Phone number
            tk.Label(
                winner_frame,
                text=f"تلفن: {self.mask_phone(phone).split()[::-1][0]}",
                font=("B Titr", font_size-4),
                fg=self.winner_fg,
                bg=self.winner_bg,
                anchor='center'
            ).pack(fill=tk.X, expand=True)

        self.previous_winners.extend(winners)
        self.save_winners_to_excel(winners)

        self.start_btn.config(state=tk.NORMAL)
        self.status_bar.config(text=f"قرعه‌کشی با موفقیت انجام شد. {len(winners)} برنده انتخاب شدند.")

    def retry_lottery(self, popup, label):
        popup.destroy()
        self.start_lottery()

    def mask_phone(self, phone):
        phone = str(phone).strip()
        
        # Remove all non-digit characters
        cleaned = ''.join(filter(str.isdigit, phone))
        
        # Validate Iranian mobile number
        if len(cleaned) != 11 or not cleaned.startswith('09'):
            return cleaned
        
        # Mask the phone number
        return f"{cleaned[7:]}***{cleaned[:4]}"

    def save_winners_to_excel(self, winners):
        try:
            file_exists = os.path.exists("winners.xlsx")
            
            if file_exists:
                wb = load_workbook("winners.xlsx")
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Winners"
                ws.append(["ردیف", "نام", "کد ملی", "شماره موبایل (مخفی)", "تاریخ قرعه‌کشی"])
                # Set column widths
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20

            start_row = ws.max_row + 1
            
            for i, (name, national_id, phone) in enumerate(winners, start=start_row):
                ws.append([
                    i-start_row+1,
                    name,
                    national_id,
                    self.mask_phone(phone),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])

            wb.save("winners.xlsx")
            self.status_bar.config(text=f"نتایج در فایل winners.xlsx ذخیره شد.")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در ذخیره فایل برندگان:\n{e}")
            self.status_bar.config(text="خطا در ذخیره نتایج")

if __name__ == "__main__":
    root = tk.Tk()
    app = LotteryApp(root)
    root.mainloop()